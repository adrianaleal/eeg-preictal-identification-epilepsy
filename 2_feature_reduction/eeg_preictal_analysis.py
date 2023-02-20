import numpy as np

import pandas as pd
import os
import math
import datetime
from pathlib import Path
import collections
# import matplotlib.pyplot as plt
import itertools

# scipy imports
from scipy.io import loadmat
from scipy.spatial.distance import pdist
from scipy.io import savemat
from scipy.spatial.distance import pdist, squareform

# sklearn imports
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
from sklearn.metrics import pairwise_distances
from sklearn.cluster import KMeans
from sklearn.cluster import AgglomerativeClustering
from sklearn.cluster import DBSCAN
from sklearn.cluster import OPTICS
from sklearn.mixture import GaussianMixture
from sklearn.metrics import silhouette_score
from sklearn.metrics import davies_bouldin_score
from sklearn.neighbors import NearestNeighbors
from sklearn.feature_selection import VarianceThreshold
from sklearn.preprocessing import MinMaxScaler

# from spherecluster import SphericalKMeans
# from soyclustering import SphericalKMeans

import hdbscan
import umap

# Excel writer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class EEGPreictalAnalysis:

    def __init__(self):
        # class initialization
        # Logger().get(display_console=True)
        # self.logger = logging.getLogger(Logger.NAME)
        self.home_dir = os.getcwd()

    @staticmethod
    def load_time_data(folder_path: str, time_filename: str) -> pd.DataFrame:
        time_info = loadmat(os.path.join(folder_path, time_filename))
        time_info = time_info[time_filename[:-4] + '_struct']
        df_date_vector = pd.DataFrame(time_info[:, 0])
        # info of data
        # df_date_vector.info()
        # print(type(df_date_vector))
        return df_date_vector

    @staticmethod
    def get_time_data(df_date_vector: pd.DataFrame, date_column: str, eeg_onset_date: str):
        # get the time vectors
        date_vector = df_date_vector[[date_column]].copy()
        if date_column == 'win_start_date':
            date_vector.loc[:, date_column] = date_vector[date_column].apply(lambda x: str(x)[1:-1])

        date_vector.loc[:, date_column] = pd.to_datetime(date_vector[date_column], infer_datetime_format=True)

        # get minute time vector
        time_delta = pd.to_datetime(eeg_onset_date, infer_datetime_format=True) - date_vector
        total_seconds = time_delta[date_column].apply(lambda x: x.seconds)
        time_vec_min = total_seconds / 60

        # datenum_vector = [self.datenum(tval) for tval in date_vector[date_column]]
        # arr_datenum = np.array(datenum_vector, dtype=np.float64)
        # time_vec_min = (arr_datenum - arr_datenum[0]) * (24 * 60)

        date_vector.loc[:, 'hour'] = date_vector[date_column].apply(lambda x: x.hour)
        date_vector.loc[:, 'minute'] = date_vector[date_column].apply(lambda x: x.minute)
        date_vector.loc[:, 'second'] = date_vector[date_column].apply(lambda x: x.second)

        h = np.array(date_vector['hour'], dtype=np.int)
        m = np.array(date_vector['minute'], dtype=np.int)
        second = np.array(date_vector['second'], dtype=np.int)

        time_vec_hour = h + m / 60 + second / 60 / 60

        df_time_minute = time_vec_min.to_frame()
        df_time_minute = df_time_minute.rename(columns={date_column: 'time_min'})
        df_time_hour = pd.DataFrame(time_vec_hour, columns=['time_hour'])

        return df_time_minute, df_time_hour

    @staticmethod
    def get_seizure_onset(patient_index: int, seizure_index: int) -> str:
        # load patient information to get seizure onset date
        patients_info = loadmat(os.path.join(os.getcwd(), 'patient_info_preictal_study_270min_before_seiz_time_gap'))
        patients_info = patients_info['patients_info_final']
        patients_name = pd.DataFrame(patients_info['patient_name'], columns=['pat_name']).apply(lambda x: x.astype(str))
        patients_name = patients_name['pat_name'].apply(lambda x: x[2:-2])
        patients_name = pd.DataFrame(patients_name, columns=['pat_name'])
        ind_pat_info = patients_name['pat_name'][patients_name['pat_name'].str.contains(str(patient_index))].index
        pat_info = pd.DataFrame(patients_info[ind_pat_info[0]])
        pat_seizure_info = pd.DataFrame(pat_info['seizures2analyseInfo'][0])
        eeg_onset_seizure = pat_seizure_info.iloc[seizure_index, 1][0]

        return eeg_onset_seizure

    @staticmethod
    def update_time_hour_night_transition(df_time_h: pd.DataFrame) -> pd.Series:
        time_vector_hour = df_time_h['time_hour']
        diff_time_hour = time_vector_hour.diff()
        if any(diff_time_hour < 0):
            # cases of midnight transition
            ind = int(diff_time_hour[diff_time_hour < -20].index.values)
            time_vec_hour_arr = time_vector_hour.to_numpy()

            # update vector of hours
            time_vec_hour_arr[ind:] = time_vec_hour_arr[ind:] + 24
            time_vector_hour_updated = pd.Series(time_vec_hour_arr)
        else:
            time_vector_hour_updated = time_vector_hour

        return time_vector_hour_updated

    def prepare_feature_data4feat_reduction(self, pat_seiz_input_folder: str, feature_group: str,
                                            indexes_5min_win2remove: pd.Series, save_dict: dict):

        # get feature folder
        feat_folder = os.path.join(pat_seiz_input_folder, 'features', feature_group.lower())
        # print(feat_folder)

        feature_file_names = os.listdir(feat_folder)

        if feature_group == 'Univariate_linear':
            # remove gamma2, gamma3 and gamma4 bands univariate linear features
            for i in range(2, 5):
                gamma_files = Path(feat_folder).glob(os.path.join(f"*[gG]amma{i}*.mat"))
                for gamma_path in gamma_files:
                    split_gamma_path = gamma_path.parts
                    if split_gamma_path[-1] in feature_file_names:
                        feature_file_names.remove(split_gamma_path[-1])

        if feature_group == 'Univariate_nonlinear':
            # remove tauPSR and eDimPSR univariate nonlinear features
            nonlinear_feat2remove = ['tauPSR.mat', 'eDimPSR.mat']
            for feat in nonlinear_feat2remove:
                if feat in feature_file_names:
                    feature_file_names.remove(feat)

        feat_files_names = pd.Series(feature_file_names)
        df_feat_data = self.load_feature_data(feat_files_names, feat_folder)

        # # inspect NaN values
        # check_nan = df_feat_data.isna().sum()
        # ind_feat_nan = check_nan.to_numpy().nonzero()

        # # count zero values
        # df_zeros = (df_feat_data == 0)
        # check_zeros = (df_zeros == 1).sum(axis=0)

        # # check if when there are NaNs there also zeros
        # get_nan = check_nan.iloc[ind_feat_nan]
        # get_zeros = check_zeros.iloc[ind_feat_nan]
        # df_get_nan_zeros = pd.concat([get_nan, get_zeros], axis=1)

        # Replace NaN elements with 0s
        df_feat_data = df_feat_data.fillna(0)

        # remove the overlapped 5-sec windows
        df_feat_data = df_feat_data.drop(df_feat_data[indexes_5min_win2remove].index)
        df_feat_data = df_feat_data.reset_index(drop=True)
        # n_feat = len(df_feat_data.columns)
        # print(f'Number of features original: {n_feat}')

        # print('row indices with Inf values')
        # df_feat_data.index[np.isinf(df_feat_data).any(1)]
        # print('columns with Inf values')
        # df_feat_data.columns.to_series()[np.isinf(df_feat_data).any()]

        # Replace Inf elements with 0s
        df_seiz_feat_data_original = df_feat_data.replace([np.inf, -np.inf], 0)

        # Removing Constant and Quasi-Constant features using Variance Threshold
        df_seiz_feat_data, constant_columns, quasi_constant_columns = self.remove_constant_features(
            df_seiz_feat_data_original)

        # DATA SCALING
        # Initialize scaler
        scaler = StandardScaler()

        # Scale each column in dataframe
        df_seiz_feat_data_scaled = pd.DataFrame(scaler.fit_transform(df_seiz_feat_data),
                                                columns=df_seiz_feat_data.columns)

        if save_dict['save_flag']:
            # save the already edited features
            pat_seiz_output_folder = save_dict['save_folder']
            feat_folder2save = os.path.join(pat_seiz_output_folder, 'features', feature_group.lower())

            if not os.path.exists(feat_folder2save):
                os.makedirs(feat_folder2save)

            len_feat = []
            for k in df_seiz_feat_data_scaled.columns:
                arr = df_seiz_feat_data_scaled[k].to_numpy(copy=True)
                len_feat.append(len(arr))
                dst = os.path.join(feat_folder2save, k + '.mat')
                savemat(dst, {'data_feature': arr}, oned_as='column')

        return df_seiz_feat_data_scaled, df_seiz_feat_data_original.columns, constant_columns, quasi_constant_columns

    def prepare_time_data4feat_reduction(self, p: int, s: int, pat_seiz_input_folder: str, save_dict: dict):

        # get time vector name
        time_file_name = 'pat_' + str(p) + '_seiz' + str(s) + '_5sec_win_info.mat'
        # print(time_file_name)

        # get time data
        df_date_vector = self.load_time_data(pat_seiz_input_folder, time_file_name)

        # remove the overlapped 5-sec windows
        indexes_5sec_overlapped = df_date_vector[['overlapped_5sec_win']]
        indexes_5sec_overlapped = indexes_5sec_overlapped['overlapped_5sec_win'].astype('bool')

        # remove the SPH 5-sec windows
        indexes_5sec_sph = df_date_vector[['SPH_5sec_win']]
        indexes_5sec_sph = indexes_5sec_sph['SPH_5sec_win'].astype('bool')

        # remove previous seizure postictal
        indexes_previous_seiz_postictal = df_date_vector[['postictal_previous_seiz']]
        indexes_previous_seiz_postictal = indexes_previous_seiz_postictal['postictal_previous_seiz'].astype('bool')

        indexes_5min_win2remove = indexes_5sec_overlapped | indexes_5sec_sph | indexes_previous_seiz_postictal

        df_date_vector = df_date_vector.drop(df_date_vector[indexes_5min_win2remove].index)

        # df_time_min, df_time_h = eeg_class.get_time_data(df_date_vector, 'win_start_date')
        if save_dict['save_flag']:
            # save the already edited time vector
            arr_rec = df_date_vector.to_records(index=False)
            arr = np.array(arr_rec).reshape((arr_rec.shape[0], 1))
            pat_seiz_output_folder = save_dict['save_folder']
            dst = os.path.join(pat_seiz_output_folder, time_file_name)
            savemat(dst, {time_file_name[:-4] + '_struct': arr}, oned_as='column')

        return df_date_vector, indexes_5min_win2remove

    @staticmethod
    def datenum(d: pd.Timestamp) -> float:
        return 366 + d.toordinal() + (d - datetime.fromordinal(d.toordinal())).total_seconds() / (24 * 60 * 60)

    @staticmethod
    def load_feature_data(feat_names: pd.Series, feat_folder: str) -> pd.DataFrame:
        # get feature values
        list_feat = []
        save_feat_names = []
        save_size = []
        # print(feat_folder)
        for f in feat_names:
            feature_val = loadmat(os.path.join(feat_folder, f))
            feat_val = feature_val['data_feature']
            list_feat.append(feat_val)
            save_feat_names.append(f[:-4])
            save_size.append([f[:-4], len(feat_val)])

        arr_seiz_feat_data = np.asarray(list_feat)
        arr_seiz_feat_data = arr_seiz_feat_data.squeeze()
        ch = arr_seiz_feat_data.shape
        # print(ch)
        if len(arr_seiz_feat_data.shape) > 2:
            arr_seiz_feat_data = arr_seiz_feat_data.reshape((arr_seiz_feat_data.shape[0]*arr_seiz_feat_data.shape[1],
                                                             arr_seiz_feat_data.shape[2]))
            electrode_file_path = Path(feat_folder).resolve().parents[1]
            electrode_file = Path(os.path.join(*list(electrode_file_path.parts))).glob('*electrode_names.mat')
            electrode_info = loadmat(str(list(electrode_file)[0]))
            electrode_names = list(electrode_info.values())[3]

            df_electrode_names = pd.DataFrame.from_dict(electrode_names).transpose()
            df_electrode_names[0] = df_electrode_names[0].str[0]

            # Extract Combination Mapping in two lists
            feat_electrode_names = list(itertools.product(save_feat_names, list(df_electrode_names[0])))
            save_feat_names = ['_'.join(words) for words in feat_electrode_names]

        df_seiz_feat_data = pd.DataFrame(arr_seiz_feat_data.transpose(), columns=save_feat_names)
        return df_seiz_feat_data

    @staticmethod
    def get_sorted_folder_content(data_dir: str, split_str: str) -> list:
        lst = os.listdir(data_dir)
        split_lst = [i.split(split_str)[1] for i in lst]
        lst = [int(x) for x in split_lst]
        lst.sort()
        return [split_str + str(i) for i in lst]

    @staticmethod
    def get_patients_names(path: str) -> list:

        pat_names = [d.rsplit('_', 1)[1] for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))]
        pat_names_sorted = sorted(pat_names, key=len)

        return pat_names_sorted

    @staticmethod
    def load_csv(filename_features_comb: str) -> pd.DataFrame:
        # print(self.filename_features_comb)
        df_files2run = pd.read_csv(filename_features_comb, delimiter=',', header=None)
        return df_files2run.rename(columns={0: 'feat_comb', 1: 'feat1', 2: 'feat2', 3: 'feat3', 4: 'pat', 5: 'seiz'})

    def load_features_to_cache(self, df_files2run: pd.DataFrame, n_feat_in_cache: int, feature_group: str) -> dict:

        df_files2run = df_files2run[['feat_comb', 'pat', 'seiz', 'feat1', 'feat2', 'feat3']]

        # get all features:
        df_files2run_all_feat = df_files2run.copy()
        df_files2run_all_feat = df_files2run_all_feat.melt(id_vars=['feat_comb', 'pat', 'seiz'], var_name='feat_index',
                                                           value_name='feat')
        df_files2run_all_feat.drop({'feat_comb', 'feat_index'}, axis='columns', inplace=True)

        # get the feature count
        df_feature_count = df_files2run_all_feat.value_counts().reset_index()
        df_feature_count = df_feature_count.rename(columns={0: 'count'})

        df_feature_count_threshold = df_feature_count.head(n_feat_in_cache)

        multi = df_feature_count_threshold.set_index(['pat', 'seiz', 'feat'])

        # store features in dictionary
        groups = multi.groupby(level=[0, 1]).indices
        feat_data_cache = {}
        for key, value in groups.items():
            pat_name = key[0]
            seiz_name = key[1]
            pat_feat_names = df_feature_count_threshold.loc[value, 'feat']

            # seizure folder path
            seiz_folder = os.path.join(self.home_dir, 'EEGFeatureSelection', pat_name, seiz_name, 'features',
                                       feature_group.lower())
            get_features = self.load_feature_data(pat_feat_names, seiz_folder)

            # dic_feat = {pat_name + '_' + seiz_name: get_features}
            feat_data_cache[pat_name + '_' + seiz_name] = []
            feat_data_cache[pat_name + '_' + seiz_name].append(get_features)

        return feat_data_cache

    @staticmethod
    def feature_redundancy(df: pd.DataFrame, cut=0.9, method='pearson'):
        # pair-wise feature redundancy computation:

        # Select the features with correlations above the threshold
        # Need to use the absolute value
        corr_mtx = df.corr(method=method).abs()

        # corr, p_val = pearsonr(df[], df[])

        # Get correlation matrix and upper triangle
        get_pairwise_corr = corr_mtx.mask(np.triu(np.ones(corr_mtx.shape)).astype(bool)).stack()
        get_pairwise_corr = get_pairwise_corr.to_frame()
        get_pairwise_corr.index = get_pairwise_corr.index.set_names(['feat1', 'feat2'])
        get_pairwise_corr = get_pairwise_corr.reset_index()
        get_pairwise_corr = get_pairwise_corr.rename(columns={0: 'coef'})

        highly_correlated_features = get_pairwise_corr[get_pairwise_corr['coef'] >= cut]
        highly_correlated_features = highly_correlated_features.reset_index(drop=True)

        # TO GET THE GRAPH OF THE CORRELATIONS --> DOES NOT WORK
        # feat_name1 = highly_correlated_features['feat1']
        # feat_name2 = highly_correlated_features['feat2']

        # Build a dataframe with your connections
        # This time a pair can appear 2 times, in one side or in the other!
        # df = pd.DataFrame({'from':feat_name1, 'to':feat_name2})

        # # Build your graph. Note that we use the Graph function to create the graph!
        # G = nx.from_pandas_edgelist(df, 'from', 'to', create_using=nx.Graph())

        # # Make the graph
        # nx.draw(G, with_labels=True, node_size=1500, alpha=0.3)
        # plt.title("UN-Directed")
        # plt.show()

        df1 = highly_correlated_features[['feat1', 'coef']]
        df2 = highly_correlated_features[['feat2', 'coef']]
        df2 = df2.rename(columns={'feat2': 'feat1'})
        df_merged = pd.concat([df1, df2])

        unique_features = df_merged['feat1'].unique()

        sum_corr_features = pd.DataFrame(0, index=unique_features, columns=['sum_corr'])

        for u in unique_features:
            ind_feat = df_merged[df_merged['feat1'] == u]
            sum_corr_features.loc[u, 'sum_corr'] = sum(ind_feat['coef'])

        sorted_features = sum_corr_features.sort_values(['sum_corr'], ascending=False)

        # count_features = df_merged['correlated_features'].value_counts()

        corr_feat = pd.DataFrame(columns=['sum_corr', 'drop_features'], index=sorted_features.index)

        drop_features = []
        for column in sorted_features.index:
            # print(column)
            # find indexes of each hi ghly correlated feature in both feat1 and feat2 columns:

            get_ind = np.where(highly_correlated_features.values == column)

            # get the sum of correlation for each feature
            get_feat_corr = highly_correlated_features['coef'][get_ind[0]]

            corr_feat.loc[column, 'sum_corr'] = get_feat_corr.sum()

            column_ind_drop_feat = []
            for i in range(len(get_ind[0])):
                row_ind = get_ind[0][i]
                col_ind = get_ind[1][i]
                if col_ind == 1:
                    feat2drop = highly_correlated_features.loc[row_ind, 'feat1']
                else:
                    feat2drop = highly_correlated_features.loc[row_ind, 'feat2']

                column_ind_drop_feat.append(feat2drop)

                if feat2drop not in drop_features and column not in drop_features:
                    drop_features.append(feat2drop)

            # get the drop features for each feature
            corr_feat.loc[column, 'drop_features'] = [', '.join(column_ind_drop_feat)]

        keep_set = df.copy()
        keep_set.drop(drop_features, axis=1, inplace=True)

        return keep_set, corr_mtx, drop_features

    @staticmethod
    def remove_constant_features(df_seiz_feat_data_original: pd.DataFrame):
        # Removing Constant and Quasi-Constant Features using Variance Threshold
        df_seiz_feat_data = df_seiz_feat_data_original.copy()
        constant_filter = VarianceThreshold(threshold=0)
        constant_filter.fit(df_seiz_feat_data)
        # unique_values = df_seiz_feat_data.columns[~constant_filter.get_support()]
        constant_columns = [column for column in df_seiz_feat_data.columns if column not in
                            df_seiz_feat_data.columns[constant_filter.get_support()]]
        df_seiz_feat_data.drop(labels=constant_columns, axis=1, inplace=True)

        # print(f'Number of features after removing constant features: {len(df_seiz_feat_data.columns)}')

        pieces = {'mean': df_seiz_feat_data.mean(), 'std': df_seiz_feat_data.std(),
                  'var': df_seiz_feat_data.var(), 'nunique': df_seiz_feat_data.nunique()}
        stats_info = pd.concat(pieces, axis=1)

        # Remove quasi-constant features:
        half_samples = int(df_seiz_feat_data.shape[0] / 2)
        dropcols_quasi_constant = stats_info[stats_info['nunique'] < half_samples]

        df_seiz_feat_data.drop(labels=dropcols_quasi_constant.index, axis=1, inplace=True)

        # print(f'Number of features after removing constant features: {len(df_seiz_feat_data.columns)}')
        return df_seiz_feat_data, constant_columns, dropcols_quasi_constant.index

    @staticmethod
    def umap_feature_reduction(df_data_scaled: pd.DataFrame, n_neighbours: int, min_dist: float):

        # print(n_neighbours, min_dist)
        reducer = umap.UMAP(random_state=42, n_components=3, n_neighbors=n_neighbours, min_dist=min_dist)
        # reducer = umap.UMAP(random_state=42, n_components=3, n_neighbors=20, min_dist=0.1)
        # default: n_neighbors=15, metric='euclidean', learning_rate=1.0
        # n_neighbors: number of approximate nearest neighbors used to construct the initial high-dimensional graph
        # min_dist: the minimum distance between points in low-dimensional space

        embedding = reducer.fit_transform(df_data_scaled)
        df_umap = pd.DataFrame(embedding, columns=['UC1', 'UC2', 'UC3'])
        # umap_parameter_tuning[(n_neighbours, min_dist)] = df_umap

        return df_umap, embedding

    @staticmethod
    def tsne_feature_reduction(df_data_scaled: pd.DataFrame, perplexity: int):
        # t-SNE
        # The algorithm is non-linear and adapts to the underlying data, performing different transformations on
        # different regions. Those differences can be a major source of confusion.
        # T-SNE with three dimensions

        # print('Perplexity', perplexity)
        tsne_3d = TSNE(random_state=42, n_components=3, perplexity=perplexity)
        # , n_iter=400, init='pca'
        # perplexity: how to balance attention between local and global aspects of data.

        # And this DataFrame contains three dimensions, built by T-SNE
        # pca_50d = PCA(n_components=50)
        # TCs_3d = tsne_3d.fit_transform(PCs_50d)
        tsne_comps_3d = tsne_3d.fit_transform(df_data_scaled)
        df_tsne = pd.DataFrame(data=tsne_comps_3d, columns=['TC1', 'TC2', 'TC3'])

        # get figure
        # fig = plt.figure(figsize=(18, 8))
        # X = np.array(df_tsne)
        # ax = plt.axes(projection='3d')
        # ax.scatter3D(X[:, 0], X[:, 1], X[:, 2], c=df_time_min)
        # plt.title(f"{tsne_selected_parameters['perplexity']}")
        # plt.show()

        return df_tsne, tsne_comps_3d

    @staticmethod
    def pca_feature_reduction(df_data_scaled: pd.DataFrame):
        # PCA
        # initialize PCA with three principal components
        pca_3d = PCA(n_components=3)

        pca_comps_3d = pca_3d.fit_transform(df_data_scaled)

        df_pca = pd.DataFrame(data=pca_comps_3d, columns=['PC1', 'PC2', 'PC3'])

        pca = pca_3d.fit(df_data_scaled)
        # print('Explained variance per principal component: {}'.format(pca.explained_variance_ratio_))
        # explain_var_ratio = pca.explained_variance_ratio_
        # plt.plot(np.cumsum(explain_var_ratio))
        # plt.xlabel('number of components')
        # plt.ylabel('cumulative explained variance')
        return df_pca, pca

    @staticmethod
    def perform_clustering(df_feat_data_scaled: pd.DataFrame, clustering_method) -> np.ndarray:

        n_clusters = int(clustering_method[-1])
        if clustering_method[:-1] == 'KMEANS_K':
            kmeans_clust = KMeans(n_clusters=n_clusters, random_state=42).fit(df_feat_data_scaled)
            clustering_solution = kmeans_clust.labels_

        elif clustering_method[:-1] == 'SPHERICAL_KMEANS_K':
            # spherical k-means
            skm = SphericalKMeans(n_clusters=n_clusters, init='k-means++')
            skm.fit(df_feat_data_scaled.to_numpy())
            clustering_solution = skm.labels_
            # skm.cluster_centers_
            # skm.inertia_

        elif clustering_method[:-1] == 'AGGLO_HIER_K':
            agglo_clust = AgglomerativeClustering(n_clusters=n_clusters).fit(df_feat_data_scaled)
            clustering_solution = agglo_clust.labels_
            # linkage: default='ward'
            # linkage is 'ward', only 'euclidean' is accepted for affinity
            # clustering_solution2 = agglo_clust.fit_predict(feat_data_scaled)
            # check = (clustering_solution == clustering_solution2).all()

        elif clustering_method[:-1] == 'HDBSCAN':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            min_cluster_size = 20  # minimum of 5 min of preictal (considering that each 5 seconds we have a sample)
            hdbscan_clust = hdbscan.HDBSCAN(min_samples=min_samples_dbscan, min_cluster_size=min_cluster_size).fit(
                df_feat_data_scaled)
            clustering_solution = hdbscan_clust.labels_

        elif clustering_method[:-1] == 'OPTICS':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            min_cluster_size = 20  # minimum of 5 min of preictal (considering that each 5 seconds we have a sample)
            optics_clust = OPTICS(min_samples=min_samples_dbscan, min_cluster_size=min_cluster_size).fit(
                df_feat_data_scaled)
            clustering_solution = optics_clust.labels_

        elif clustering_method == 'DBSCAN_D1':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            dbscan_clust = DBSCAN(eps=0.25, min_samples=min_samples_dbscan).fit(df_feat_data_scaled)
            clustering_solution = dbscan_clust.labels_
            # Noisy samples are given the label -1.
            # clustering_solution2 = dbscan_clust.fit_predict(feat_data_scaled)
            # check = (clustering_solution == clustering_solution2).all()

        elif clustering_method == 'DBSCAN_D2':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            dbscan_clust = DBSCAN(eps=0.35, min_samples=min_samples_dbscan).fit(df_feat_data_scaled)
            clustering_solution = dbscan_clust.labels_

        elif clustering_method == 'DBSCAN_D3':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            dbscan_clust = DBSCAN(eps=0.45, min_samples=min_samples_dbscan).fit(df_feat_data_scaled)
            clustering_solution = dbscan_clust.labels_

        elif clustering_method == 'DBSCAN_D4':
            dimension = len(df_feat_data_scaled.columns)
            min_samples_dbscan = 2 * dimension
            dbscan_clust = DBSCAN(eps=0.55, min_samples=min_samples_dbscan).fit(df_feat_data_scaled)
            clustering_solution = dbscan_clust.labels_

        elif clustering_method[:-1] == 'GMM_K':
            lowest_bic = np.infty
            best_gmm = None
            bic = []
            cv_types = ['spherical', 'tied', 'diag', 'full']
            for cv_type in cv_types:
                # Fit a Gaussian mixture with EM
                gmm = GaussianMixture(n_components=n_clusters, random_state=42, max_iter=500, covariance_type=cv_type,
                                      tol=1e-2)
                gmm.fit(df_feat_data_scaled)
                bic.append(gmm.bic(df_feat_data_scaled))
                if bic[-1] < lowest_bic:
                    lowest_bic = bic[-1]
                    best_gmm = gmm
            # print(best_gmm.covariance_type)
            clustering_solution = best_gmm.predict(df_feat_data_scaled)
            # reg_covar: default=1e-6
            # max_iter: default=100
            # init_params: {'kmeans', 'random'}, default='kmeans'
            # Plot the BIC scores
            # plt.figure(figsize=(8, 6))
            # plt.bar(cv_types, bic)
            # plt.ylim([bic.min() * 1.01 - .01 * bic.max(), bic.max()])
            # plt.title('BIC score per model')
            # plt.xlabel('Covariance type', fontsize=14)
            # plt.ylabel('Square of Value', fontsize=14)
            # a = range(0, len(cv_types))
            # plt.xticks(a, cv_types, rotation='vertical')
            # plt.show()

        else:
            clustering_solution = None
        return clustering_solution

    @staticmethod
    def assess_frequency_clusters(clustering_solution: np.ndarray):
        frequency_clusters = collections.Counter(clustering_solution)
        df_frequency_clusters = pd.DataFrame.from_dict(frequency_clusters, orient='index').reset_index()
        df_frequency_clusters = df_frequency_clusters.rename(columns={'index': 'value', 0: 'count'}, inplace=False)
        # print(df_frequency_clusters)

        clustering_solution_changed = clustering_solution.copy()
        df_frequency_clusters_changed = df_frequency_clusters

        if df_frequency_clusters.loc[0, 'count'] == df_frequency_clusters.loc[1, 'count']:
            # when the two clusters have the same amount of samples assign class 2 to the cluster with the beginning
            # sample that is closest to the seizure:
            mask1 = clustering_solution == df_frequency_clusters.loc[0, 'value']
            mask2 = clustering_solution == df_frequency_clusters.loc[1, 'value']
            indexes_cluster11 = [i for i, val in enumerate(mask1) if val]
            indexes_cluster22 = [i for i, val in enumerate(mask2) if val]

            # indexes_cluster1 = np.where(clustering_solution == df_frequency_clusters.loc[0, 'value'])[0]
            # indexes_cluster2 = np.where(clustering_solution == df_frequency_clusters.loc[1, 'value'])[0]

            if indexes_cluster11[0] > indexes_cluster22[0]:
                clustering_solution_changed[mask1] = df_frequency_clusters.loc[1, 'value']
                clustering_solution_changed[mask2] = df_frequency_clusters.loc[0, 'value']
        else:
            # assign class 1 with the highest amount of samples if that is not already the case:
            # df_frequency_clusters_sorted.sort_values(['value', 'count'], ascending=[True, True], inplace=True)
            df_frequency_clusters_sorted = pd.DataFrame(np.sort(df_frequency_clusters.values, axis=0),
                                                        index=df_frequency_clusters.index,
                                                        columns=df_frequency_clusters.columns)

            if df_frequency_clusters_sorted.equals(df_frequency_clusters):
                # change the cluster classes
                df_frequency_clusters_sorted = df_frequency_clusters_sorted.sort_values(['count'], ascending=False,
                                                                                        ignore_index=True)
                for c in range(0, df_frequency_clusters_sorted.shape[0]):
                    clustering_solution_changed[clustering_solution == df_frequency_clusters_sorted.loc[c, 'value']] = \
                        df_frequency_clusters_sorted['value'].iloc[-1 - c]

                frequency_clusters_changed = collections.Counter(clustering_solution_changed)
                df_frequency_clusters_changed = pd.DataFrame.from_dict(frequency_clusters_changed, orient='index'). \
                    reset_index()
                df_frequency_clusters_changed = df_frequency_clusters_changed.rename(
                    columns={'index': 'value', 0: 'count'},
                    inplace=False)
                df_frequency_clusters_changed = df_frequency_clusters_changed.sort_values('value', ignore_index=True)

        return clustering_solution_changed, df_frequency_clusters_changed

    @staticmethod
    def centroid_clust_eval(data2cluster: pd.DataFrame, clustering_solution: np.ndarray):
        clusters = set(clustering_solution)
        n_clusters = len(clusters)
        # print(n_clusters)

        dim = data2cluster.shape[1]
        deviations = np.empty((n_clusters, 1))
        deviations[:] = np.nan

        cluster_centroids = np.empty((n_clusters, dim))
        for kk, val in enumerate(clusters):
            cluster = data2cluster.iloc[clustering_solution == val, :]
            n_samples = cluster.shape[0]
            cluster = cluster.to_numpy()

            if n_samples > 1:
                centroid = cluster.mean(axis=0)
            else:
                centroid = cluster

            cluster_centroids[kk, :] = centroid
            centroid = centroid.reshape((1, dim))
            dist = pairwise_distances(cluster, centroid)

            deviations[kk] = sum(dist)

        overall_deviation = deviations.sum()

        if n_clusters > 1:
            get_dist = pdist(cluster_centroids)
            cluster_separation = 2 / (n_clusters * (n_clusters - 1)) * sum(get_dist)
        else:
            cluster_separation = None

        return overall_deviation, cluster_separation

    @staticmethod
    def connectivity(data2cluster: pd.DataFrame, clustering_solution: np.ndarray):
        n_neighbors = 8
        nbrs = NearestNeighbors(n_neighbors=n_neighbors, metric='euclidean').fit(data2cluster)
        distances, indices = nbrs.kneighbors(data2cluster)
        indices = indices.transpose()

        clusters = set(clustering_solution)
        n_clusters = len(clusters)

        n_samples = data2cluster.shape[0]
        connection_matrix = np.zeros((n_samples, n_neighbors))

        for cc in range(0, n_samples - 1):
            for ind_j_in_nn_matrix in range(1, n_neighbors):
                count = 0
                i_val = clustering_solution[cc]
                j_val = clustering_solution[indices[ind_j_in_nn_matrix, cc]]

                for kk in range(n_clusters):
                    if i_val == kk and j_val == kk:
                        count = count + 1

                if count == 0:
                    connection_matrix[cc, ind_j_in_nn_matrix - 1] = 1 / ind_j_in_nn_matrix

        conn = sum(sum(connection_matrix))

        return conn

    @staticmethod
    def get_clust_pairs(clusters) -> list:
        # from https://github.com/crew102/validclust/blob/master/validclust/indices.py
        return [(i, j) for i in clusters for j in clusters if i > j]

    def dunn(self, clustering_solution: np.ndarray, dist_mat: np.ndarray) -> float:
        # from https://github.com/crew102/validclust/blob/master/validclust/indices.py
        clusters = set(clustering_solution)
        inter_dists = [dist_mat[np.ix_(clustering_solution == i, clustering_solution == j)].min() for i, j in
                       self.get_clust_pairs(clusters)]
        intra_dists = [dist_mat[np.ix_(clustering_solution == i, clustering_solution == i)].max() for i in clusters]

        return min(inter_dists) / max(intra_dists)

    def dunn_2clusters(self, clustering_solution: np.ndarray, dist_mat: np.ndarray) -> float:
        clusters = set(clustering_solution)
        clust_val = self.get_clust_pairs(clusters)
        inter_dists = [
            dist_mat[np.ix_(clustering_solution == clust_val[0][0], clustering_solution == clust_val[0][1])].min()]

        intra_dists = [
            dist_mat[np.ix_(clustering_solution == clust_val[0][0], clustering_solution == clust_val[0][0])].max(),
            dist_mat[np.ix_(clustering_solution == clust_val[0][1], clustering_solution == clust_val[0][1])].max()]

        return min(inter_dists) / max(intra_dists)

    def cluster_evaluation_indexes(self, df_data: pd.DataFrame, clustering_solution: np.ndarray,
                                   distances_matrix: np.ndarray):
        n_cluster_values = np.unique(clustering_solution)
        n_clusters = len(n_cluster_values)
        if np.all(clustering_solution):
            n_samples_smaller_cluster = len(clustering_solution)
        else:
            samples_count_in_cluster = np.bincount(clustering_solution[clustering_solution >= 0])
            n_samples_smaller_cluster = np.min(samples_count_in_cluster)
        noisy_clusters = any(n_cluster_values == -1)
        if n_clusters > 1:
            dunns_index = self.dunn(clustering_solution, distances_matrix)

            # Silhouette index
            si_score = silhouette_score(df_data, clustering_solution, metric='euclidean')

            # overall deviation
            overall_deviation, cluster_separation = self.centroid_clust_eval(df_data, clustering_solution)

            # connectivity
            conn = self.connectivity(df_data, clustering_solution)

            # Davies-Bouldin Index
            dbi_score = davies_bouldin_score(df_data, clustering_solution)
        else:
            dunns_index = None
            si_score = None
            overall_deviation = None
            cluster_separation = None
            conn = None
            dbi_score = None
        return [n_clusters, noisy_clusters, n_samples_smaller_cluster, dunns_index, si_score, overall_deviation,
                cluster_separation, conn, dbi_score]

    def perform_cluster_solution_evaluation(self, df_feat_data_scaled: pd.DataFrame, clustering_solution: np.ndarray,
                                            distances_matrix: np.ndarray, df_clust_eval_results: pd.DataFrame,
                                            save_clust_eval_folder: str, feature_comb: int, clust_method):

        clustering_solution, df_frequency_clusters = self.assess_frequency_clusters(clustering_solution)

        n_samples_smaller_cluster = df_frequency_clusters['count'].iloc[-1]
        df_clust_eval_results.iloc[6, 1] = n_samples_smaller_cluster

        min_5min_win_samples = (5 * 60) / 5

        if n_samples_smaller_cluster >= min_5min_win_samples:

            dunns_index = self.dunn(clustering_solution, distances_matrix)

            df_clust_eval_results.iloc[7, 1] = dunns_index
            if dunns_index >= 0.15:

                # Silhouette index
                si_score = silhouette_score(df_feat_data_scaled, clustering_solution, metric='euclidean')
                df_clust_eval_results.iloc[8, 1] = si_score

                # overall deviation
                overall_deviation, cluster_separation = self.centroid_clust_eval(df_feat_data_scaled,
                                                                                 clustering_solution)
                df_clust_eval_results.iloc[9, 1] = overall_deviation
                df_clust_eval_results.iloc[10, 1] = cluster_separation

                # connectivity
                conn = self.connectivity(df_feat_data_scaled, clustering_solution)
                df_clust_eval_results.iloc[11, 1] = conn

                # Davies-Bouldin Index
                dbi_score = davies_bouldin_score(df_feat_data_scaled, clustering_solution)
                df_clust_eval_results.iloc[12, 1] = dbi_score

                # Evaluate time continuity
                indexes_smaller_cluster = np.where(clustering_solution == df_frequency_clusters['value'].iloc[-1])
                diff_vec = np.diff(indexes_smaller_cluster) == 1
                df_clust_eval_results.iloc[13, 1] = diff_vec.all()

                # Start and end samples of the smaller cluster
                # sample_start_smaller_cluster:
                df_clust_eval_results.iloc[14, 1] = indexes_smaller_cluster[0][1]

                # sample_end_smaller_cluster:
                df_clust_eval_results.iloc[15, 1] = indexes_smaller_cluster[0][-1]

                if not os.path.isdir(save_clust_eval_folder):
                    os.makedirs(save_clust_eval_folder)

                df_clust_eval_results.to_csv(os.path.join(save_clust_eval_folder, 'feat_comb' + str(feature_comb) + '_'
                                                          + clust_method.name.lower() + '.csv'), index=False)

        return

    def run_feat_comb(self, df_feat_data: pd.DataFrame, feat_comb: int, feat_names: list, pat_name: str, seiz_name: str,
                      plot_figure: bool, feature_group: str):

        # seizure folder path
        save_clust_eval_folder = os.path.join(self.home_dir, 'EEGFeatureSelection', pat_name, seiz_name,
                                              'clustering_' + feature_group.lower())

        # DATA SCALING
        # Initialize scaler
        scaler = StandardScaler()

        # Scale each column in dataframe
        df_feat_data_scaled = pd.DataFrame(scaler.fit_transform(df_feat_data))

        # get pairwise distances between samples
        distances_matrix = pairwise_distances(df_feat_data_scaled)

        clustering_methods = ['KMEANS_K2', 'KMEANS_K3', 'KMEANS_K4', 'AGGLO_HIER_K2', 'AGGLO_HIER_K3', 'AGGLO_HIER_K4',
                              'HDBSCAN', 'GMM_K2', 'GMM_K3', 'GMM_K4']
        # DBSCAN_D1, DBSCAN_D2, DBSCAN_D3, DBSCAN_D4

        first_col = ['feat_comb', 'feat1', 'feat2', 'feat3', 'n_clusters', 'noisy_clusters',
                     'n_samples_smaller_cluster', 'DI', 'SI', 'OD', 'CS', 'C', 'DBI', 'continuity_smaller_cluster',
                     'sample_start_smaller_cluster', 'sample_end_smaller_cluster']

        # start_time = datetime.now()
        for cm in clustering_methods:
            clustering_solution = self.perform_clustering(df_feat_data_scaled, cm)

            n_cluster_values = np.unique(clustering_solution)

            df_clust_eval_results = pd.DataFrame(columns=['clustering_evaluation_metrics', cm.lower()])
            df_clust_eval_results.clustering_evaluation_metrics = first_col
            df_clust_eval_results.iloc[0, 1] = feat_comb
            df_clust_eval_results.iloc[1, 1] = feat_names[0]
            df_clust_eval_results.iloc[2, 1] = feat_names[1]
            df_clust_eval_results.iloc[3, 1] = feat_names[2]
            df_clust_eval_results.iloc[4, 1] = len(n_cluster_values)
            df_clust_eval_results.iloc[5, 1] = any(n_cluster_values == -1)

            if len(n_cluster_values) == 2 and not any(n_cluster_values == -1):
                self.perform_cluster_solution_evaluation(df_feat_data_scaled, clustering_solution, distances_matrix,
                                                         df_clust_eval_results, save_clust_eval_folder, feat_comb, cm)

            if plot_figure:
                time_folder = os.path.join(self.home_dir, 'EEGFeatureSelection', pat_name, seiz_name)
                time_file_name = pat_name + '_seiz' + seiz_name[-1] + '_5sec_win_info.mat'
                df_date_vector = self.load_time_data(time_folder, time_file_name)
                pat_seiz_name = pat_name + '_seiz' + seiz_name[-1]
                df_time_min, df_time_h = self.get_time_data(df_date_vector, 'win_start_date')
                self.plot_cluster_solution(cm, feat_names, pat_seiz_name, df_feat_data_scaled, clustering_solution,
                                           df_time_min)

        return

    @staticmethod
    def select_parameter_feat_reduction(results_folder: str, parameter_list: list, reduct_model: str,
                                        clust_eval_index: str, min_cluster_size: int):
        # choose the parameter in list for a given feature reduction model
        # print(results_folder)
        # load the file
        df_clust_results = pd.read_csv(os.path.join(results_folder, 'select_' + reduct_model + '.csv'), delimiter=',')

        # find maximum of a given cluster evaluation index excluding the cluster with noisy samples from hdbscan
        # ind_noisy_clusters = df_clust_results['noisy_clusters']
        ind_n_samples_smaller_cluster = df_clust_results['n_samples_smaller_cluster']
        # min_cluster_size = 20  # minimum of 5 min of preictal (considering that each 5 seconds we have a sample)

        df_clust_results_over_min_cluster_size = df_clust_results.loc[ind_n_samples_smaller_cluster >= min_cluster_size, :]
        max_clust_eval_index = df_clust_results_over_min_cluster_size[clust_eval_index].max()
        ind_max_clust_eval_index = df_clust_results_over_min_cluster_size[clust_eval_index] == max_clust_eval_index
        df_max_clust_eval_index = df_clust_results_over_min_cluster_size.loc[ind_max_clust_eval_index, :]
        if len(df_max_clust_eval_index) > 1:
            df_max_clust_eval_index_no_noisy = df_max_clust_eval_index.loc[~df_max_clust_eval_index['noisy_clusters'], :]
            if len(df_max_clust_eval_index_no_noisy) > 0:
                ind_final_max_clust_eval_index = df_max_clust_eval_index_no_noisy.index
            else:
                ind_final_max_clust_eval_index = df_max_clust_eval_index.index
        else:
            ind_final_max_clust_eval_index = df_max_clust_eval_index.index

        parameter_values = []
        for i in parameter_list:
            get_parameter = df_clust_results.loc[ind_final_max_clust_eval_index[0], i]
            parameter_values.append(get_parameter)
        dict_out = dict(zip(parameter_list, parameter_values))
        return dict_out, df_clust_results.loc[ind_final_max_clust_eval_index[0], clust_eval_index]

    @staticmethod
    def plot_cluster_solution(cm, feat_names, pat_seiz_name, df_feat_data_scaled, clustering_solution, time_vec_min):
        # Plot the clustering solution
        import matplotlib.pyplot as plt

        # set up a figure twice as wide as it is tall
        fig = plt.figure(figsize=(14, 8))
        main_title = cm.replace('_', ' ') + '_' + pat_seiz_name.replace('_', ' ')
        fig.suptitle(main_title)

        feat_data_scaled = np.array(df_feat_data_scaled)

        # ===============
        #  First subplot
        # ===============
        # set up the axes for the first plot
        ax = fig.add_subplot(1, 2, 1, projection='3d')
        ax.scatter3D(feat_data_scaled[:, 0], feat_data_scaled[:, 1], feat_data_scaled[:, 2], c=clustering_solution)
        # plt.title('clustering solution')
        plt.subplots_adjust(hspace=.35, bottom=.02)
        ax.set_xlabel(feat_names[0][0:-4].replace('_', ' '))
        ax.set_ylabel(feat_names[1][0:-4].replace('_', ' '))
        ax.set_zlabel(feat_names[2][0:-4].replace('_', ' '))

        # ===============
        # Second subplot
        # ===============
        # set up the axes for the second plot
        ax = fig.add_subplot(1, 2, 2, projection='3d')
        sctt2 = ax.scatter3D(feat_data_scaled[:, 0], feat_data_scaled[:, 1], feat_data_scaled[:, 2], c=time_vec_min)
        # plt.title('data over time')
        plt.subplots_adjust(hspace=.35, bottom=.02)
        ax.set_xlabel(feat_names[0][0:-4].replace('_', ' '))
        ax.set_ylabel(feat_names[1][0:-4].replace('_', ' '))
        ax.set_zlabel(feat_names[2][0:-4].replace('_', ' '))
        cbar_ax = fig.add_axes([0.95, 0.1, 0.01, 0.7])
        fig.colorbar(sctt2, cax=cbar_ax)
        cbar_ax.set_label('time (min)')
        plt.show()
        return

    @staticmethod
    def get_time_information(df_dim_reduct: pd.DataFrame, eeg_onset_date):

        # get the two-hour interval to select the preictal interval from
        eeg_onset_date_timestamp = pd.to_datetime(eeg_onset_date, infer_datetime_format=True)
        two_hour_date = eeg_onset_date_timestamp - datetime.timedelta(seconds=2 * 60 * 60)
        two_hour_interval_hour = two_hour_date.hour + two_hour_date.minute / 60 + two_hour_date.second / 60 / 60

        time_vector_hour = df_dim_reduct['time_hour']
        time_vec_minutes = df_dim_reduct['time_min']
        two_hour_interval_min = 120  # time_vec_minutes.iloc[-1] +

        diff_time_hour = time_vector_hour.diff()
        if any(diff_time_hour < 0):
            # cases of midnight transition
            ind = int(diff_time_hour[diff_time_hour < -20].index.values)
            time_vec_hour_arr = time_vector_hour.to_numpy()

            # update two hour interval
            ind_two_hour = time_vec_hour_arr[ind:] >= two_hour_interval_hour
            if not ind_two_hour[0] and len(np.unique(ind_two_hour)) > 1:
                two_hour_interval_hour = two_hour_interval_hour + 24

            # update vector of hours
            time_vec_hour_arr[ind:] = time_vec_hour_arr[ind:] + 24
            time_vector_hour = pd.Series(time_vec_hour_arr)
        return time_vector_hour, time_vec_minutes, two_hour_interval_hour, two_hour_interval_min

    @staticmethod
    def get_sleep_vector(df_hypnogram: pd.DataFrame, time_vec_hour: pd.Series, eeg_onset_date: str,
                         sph_date: pd.Timestamp, eeg_class_obj, clustering_solution: np.ndarray):

        # remove the SPH from the hypnogram
        df_hypnogram_datetime = pd.to_datetime(df_hypnogram['time_date'])
        indexes_sph2remove = df_hypnogram_datetime > sph_date
        df_hypnogram = df_hypnogram.drop(df_hypnogram[indexes_sph2remove].index)
        df_time_min, df_time_h = eeg_class_obj.get_time_data(df_hypnogram, 'time_date', eeg_onset_date)
        df_hypnogram = pd.concat([df_hypnogram, df_time_h], axis=1)

        diff_time_hour = df_hypnogram['time_hour'].diff()
        if any(diff_time_hour < 0):
            # cases of midnight transition
            ind = int(diff_time_hour[diff_time_hour < -20].index.values)
            time_vec_hour_arr = df_hypnogram['time_hour'].to_numpy()
            time_vec_hour_arr[ind:] = time_vec_hour_arr[ind:] + 24
            df_hypnogram['time_hour'] = time_vec_hour_arr

        sleep_vec = np.zeros((len(clustering_solution),), dtype=int)
        for i in range(len(df_hypnogram)):
            if df_hypnogram.loc[i, 'sleep_stage'] == 1:
                if i + 1 > len(df_hypnogram) - 1:
                    indexes = time_vec_hour >= df_hypnogram.loc[i, 'time_hour']
                else:
                    indexes = (time_vec_hour >= df_hypnogram.loc[i, 'time_hour']) & \
                              (time_vec_hour < df_hypnogram.loc[i + 1, 'time_hour'])
                sleep_vec[indexes] = 1

        return sleep_vec

    @staticmethod
    def get_sleep_vector_control(df_hypnogram: pd.DataFrame, time_vec_hour: pd.Series, last_date: str,
                                 df_control_datetimes: pd.DataFrame, eeg_class_obj, clustering_solution: np.ndarray):

        # get sleep vector within the control datetime interval
        start_control_datetime = df_control_datetimes.loc[0, 0]
        end_control_datetime = df_control_datetimes.loc[1, 0]

        df_hypnogram_datetime = df_hypnogram[['time_date']]

        # get control intervals start and end indexes
        control_intervals_indexes = (df_hypnogram_datetime >= start_control_datetime) & (
                    df_hypnogram_datetime <= end_control_datetime)
        idx_control_start = control_intervals_indexes.idxmax()[0]
        idx_control_end = control_intervals_indexes.where(control_intervals_indexes).last_valid_index()

        # get the control interval datetimes vector
        df_hypnogram_control = df_hypnogram.copy()
        df_hypnogram_control = df_hypnogram_control[idx_control_start:idx_control_end]
        df_hypnogram_control = df_hypnogram_control.reset_index(drop=True)
        # remove the samples outside control datetimes

        df_time_min, df_time_h = eeg_class_obj.get_time_data(df_hypnogram_control, 'time_date', last_date)
        df_hypnogram_control = pd.concat([df_hypnogram_control, df_time_h], axis=1)

        diff_time_hour = df_hypnogram_control['time_hour'].diff()
        if any(diff_time_hour < 0):
            # cases of midnight transition
            ind = int(diff_time_hour[diff_time_hour < -20].index.values)
            time_vec_hour_arr = df_hypnogram_control['time_hour'].to_numpy()
            time_vec_hour_arr[ind:] = time_vec_hour_arr[ind:] + 24
            df_hypnogram_control['time_hour'] = time_vec_hour_arr

        sleep_vec = np.zeros((len(clustering_solution),), dtype=int)
        for i in range(len(df_hypnogram_control)):
            if df_hypnogram_control.loc[i, 'sleep_stage'] == 1:
                if i + 1 > len(df_hypnogram_control) - 1:
                    indexes = time_vec_hour >= df_hypnogram_control.loc[i, 'time_hour']
                else:
                    indexes = (time_vec_hour >= df_hypnogram_control.loc[i, 'time_hour']) & \
                              (time_vec_hour < df_hypnogram_control.loc[i + 1, 'time_hour'])
                sleep_vec[indexes] = 1

        return sleep_vec

    @staticmethod
    def sort_clustering_solution(clustering_solution: np.ndarray):
        # sort the clustering solution according to the highest number of samples

        if np.any(clustering_solution < 0):
            clustering_solution = clustering_solution + 1

        frequency_clusters = collections.Counter(clustering_solution)
        df_frequency_clusters = pd.DataFrame.from_dict(frequency_clusters, orient='index').reset_index()
        df_frequency_clusters = df_frequency_clusters.rename(columns={'index': 'value', 0: 'count'}, inplace=False)
        # print(df_frequency_clusters)
        df_frequency_clusters_sorted = pd.concat([df_frequency_clusters['value'].sort_values(ignore_index=True,
                                                                                             ascending=True),
                                                  df_frequency_clusters['count'].sort_values(ignore_index=True,
                                                                                             ascending=False)], axis=1)

        if not df_frequency_clusters.equals(df_frequency_clusters_sorted):
            # reassign samples in clustering solution
            clustering_solution_changed = np.empty(clustering_solution.shape, dtype=int)
            for a in np.unique(df_frequency_clusters_sorted['count']):
                ind_count = np.where((df_frequency_clusters['count'] == a) == True)
                ind_count_sorted = np.where((df_frequency_clusters_sorted['count'] == a) == True)

                for pair in zip(ind_count_sorted[0], ind_count[0]):
                    new_value = df_frequency_clusters_sorted.loc[pair[0], 'value']
                    old_value = df_frequency_clusters.loc[pair[1], 'value']
                    # print('iter')
                    # print([old_value, new_value])
                    clustering_solution_changed[clustering_solution == old_value] = new_value

            # check = pd.concat([pd.DataFrame(clustering_solution), pd.DataFrame(clustering_solution_changed)], axis=1)
            check_solution = np.unique(np.add(clustering_solution, clustering_solution_changed))
            if len(check_solution) != 1:
                print('PROBLEM WITH CLUSTERING SOLUTION')
        else:
            clustering_solution_changed = clustering_solution

        return clustering_solution_changed, df_frequency_clusters, df_frequency_clusters_sorted

    @staticmethod
    def get_time_connectivity(n_neigh: int, df_dim_reduct_input: pd.DataFrame, time_vec_minute: pd.Series,
                              time_vec_hour: pd.Series):
        # normalize data between [-1 1]
        scaler = MinMaxScaler(feature_range=(-1, 1))
        scaler.fit(df_dim_reduct_input)
        df_dim_reduct = pd.DataFrame(scaler.transform(df_dim_reduct_input))

        dist_upper_tri = pdist(df_dim_reduct.to_numpy())
        dist_mat = squareform(dist_upper_tri)

        proxindex = 1 - (sum(np.diag(dist_mat, 1)) / ((len(df_dim_reduct) - 1) * 2 * np.sqrt(3)))
        # proxindex = 1 - (sum(dist_upper_tri) / (len(dist_upper_tri) * 2 * np.sqrt(3)))

        return proxindex

    def cluster_time_connectivity(self, n_neigh, n_clust_val, df_dim_reduct, clustering_solution, time_vec_minute,
                                  time_vec_hour):
        cluster_time_connect = []
        for c in n_clust_val:
            indexes_cluster = clustering_solution == c
            df_dim_reduction_cluster = df_dim_reduct.iloc[indexes_cluster, :]
            time_vec_min_cluster = time_vec_minute.iloc[indexes_cluster]
            time_vec_min_cluster.index = range(len(time_vec_min_cluster))
            time_vec_h_cluster = time_vec_hour.iloc[indexes_cluster]
            time_vec_h_cluster.index = range(len(time_vec_h_cluster))
            cluster_connect = self.get_time_connectivity(n_neigh, df_dim_reduction_cluster, time_vec_min_cluster,
                                                         time_vec_h_cluster)
            cluster_time_connect.append(cluster_connect)
        return cluster_time_connect

    @staticmethod
    def get_phi_coefficient(clustering_solution: np.ndarray, sleep_vector_best_model_paper: np.ndarray):

        if (len(np.unique(sleep_vector_best_model_paper)) == 2) & (len(np.unique(clustering_solution)) == 2):
            # compute the Phi coefficient:
            sum_vectors = clustering_solution + sleep_vector_best_model_paper
            diff_vectors = clustering_solution - sleep_vector_best_model_paper
            a = np.int64((sum_vectors == 2).sum())
            d = np.int64((sum_vectors == 0).sum())
            b = np.int64((diff_vectors == 1).sum())
            c = np.int64((diff_vectors == -1).sum())
            # print(a, b, c, d)
            peirce_I = round(abs((a * d - b * c) / ((a + b) * (c + d))), 2)
            phi_coefficient = round(abs((a * d - b * c) / math.sqrt((a + b) * (a + c) * (b + d) * (c + d))), 2)
        else:
            phi_coefficient = None
            peirce_I = None

        return phi_coefficient, peirce_I

    @staticmethod
    def save_to_excel_columns(group_features: str, df_ind_pat_seiz_class: pd.DataFrame, excel_filename: str):
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb[group_features.lower()]
        for index, row in df_ind_pat_seiz_class.iterrows():
            print(index)
            print(row)
            cell = 'A%d' % (index + 2)
            ws[cell] = row[0]
            cell = 'B%d' % (index + 2)
            ws[cell] = row[1]
        wb.save('results_class3.xlsx')
        return

    @staticmethod
    def save_to_excel_file(row_name2write: str, worksheet_name: str, headers, values2write: list, file_path: str,
                           file_name: str):

        # path where results file is stored

        # If path doesn't exist, create it
        os.makedirs(file_path, exist_ok=True)

        # Define filename and path to file
        file_path_name = os.path.join(file_path, file_name + '.xlsx')

        # If file exists, open it
        if os.path.exists(file_path_name):

            # Load Workbook object
            wb = openpyxl.load_workbook(file_path_name)

            # Get active sheet
            # ws = wb.active
            if worksheet_name not in wb.sheetnames:
                wb.create_sheet(worksheet_name)
                ws = wb[worksheet_name]
                # Write headers
                for col, val in enumerate(headers, start=1):
                    ws.cell(row=1, column=col).value = val
                    ws.cell(row=1, column=col).font = Font(color='00003366', bold=True)
                    ws.cell(row=1, column=col).fill = PatternFill('solid', fgColor='0099CCFF')
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.column_dimensions[get_column_letter(col)].width = max(len(val), 10)

                # Get row number to write
                row = ws.max_row + 1

            else:
                ws = wb[worksheet_name]
                # Get row number to write
                df = pd.read_excel(file_path_name, sheet_name=worksheet_name)
                # it checks the first column:
                get_pat_ind = df.loc[df[df.columns[0]] == row_name2write].index.values
                if get_pat_ind.size == 0:
                    row = ws.max_row + 1
                else:
                    row = get_pat_ind[0] + 2

        # Else, create new file
        else:

            # Create new Workbook object
            wb = openpyxl.Workbook()

            # Get active sheet
            ws = wb.active
            ws.title = worksheet_name

            # Write headers
            for col_ind, col_val in enumerate(headers, start=1):
                ws.cell(row=1, column=col_ind).value = col_val
                ws.cell(row=1, column=col_ind).font = Font(color='00003366', bold=True)
                ws.cell(row=1, column=col_ind).fill = PatternFill('solid', fgColor='0099CCFF')
                ws.cell(row=1, column=col_ind).alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col_ind)].width = max(len(col_val), 10)

            # Get row number to write
            row = ws.max_row + 1

        # Write data into file

        # Write patient results
        for col, val in enumerate(values2write, start=1):
            ws.cell(row=row, column=col).value = val
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        # Save file
        wb.save(file_path_name)
        wb.close()


if __name__ == '__main__':

    eeg_class = EEGPreictalAnalysis()
    # print('end')
